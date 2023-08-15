import pygame
import random
import sys
import openpyxl
from openpyxl import Workbook

# Initialize pygame
pygame.init()

# Constants
SCREEN_WIDTH = 640
SCREEN_HEIGHT = 480
CELL_SIZE = 20
BORDER_SIZE = 20
WHITE = (255, 255, 255)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
BLACK = (0, 0, 0)
FONT = pygame.font.Font(None, 36)


def setup():
    pygame.init()
    return pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))


# Base classes
class Snake:
    def __init__(self):
        self.positions = [(100, 100), (80, 100), (60, 100)]
        self.direction = (CELL_SIZE, 0)

    def update(self):
        self.positions.insert(
            0,
            (
                self.positions[0][0] + self.direction[0],
                self.positions[0][1] + self.direction[1],
            ),
        )
        self.positions.pop()

    def handle_keys(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_UP and self.direction != (0, CELL_SIZE):
                    self.direction = (0, -CELL_SIZE)
                elif event.key == pygame.K_DOWN and self.direction != (0, -CELL_SIZE):
                    self.direction = (0, CELL_SIZE)
                elif event.key == pygame.K_LEFT and self.direction != (CELL_SIZE, 0):
                    self.direction = (-CELL_SIZE, 0)
                elif event.key == pygame.K_RIGHT and self.direction != (-CELL_SIZE, 0):
                    self.direction = (CELL_SIZE, 0)


class Food:
    def __init__(self, obstacles):
        self.color = RED
        self.obstacles = obstacles
        self.snake_positions = []  # Store snake positions for collision check
        self.update()

    def update(self):
        valid_positions = [
            (x, y)
            for x in range(BORDER_SIZE, SCREEN_WIDTH - BORDER_SIZE, CELL_SIZE)
            for y in range(BORDER_SIZE, SCREEN_HEIGHT - BORDER_SIZE, CELL_SIZE)
            if (x, y) not in [obstacle.position for obstacle in self.obstacles]
            and (x, y) not in self.snake_positions
        ]

        self.position = random.choice(valid_positions)


# Obstacle class
class Obstacle:
    def __init__(self):
        self.color = BLACK
        self.update()

    def update(self):
        valid_positions = [
            (x, y)
            for x in range(BORDER_SIZE, SCREEN_WIDTH - BORDER_SIZE, CELL_SIZE)
            for y in range(BORDER_SIZE, SCREEN_HEIGHT - BORDER_SIZE, CELL_SIZE)
        ]
        self.position = random.choice(valid_positions)


def game_over(screen, score):
    screen.fill(WHITE)
    game_over_text = FONT.render("Game Over", True, RED)
    final_score_text = FONT.render(f"Final Score: {score}", True, RED)
    save_prompt_text = FONT.render("Do you want to save your result?", True, BLACK)
    message_gap = 40  # Gap between the message and buttons
    button_gap = 10  # Gap between buttons

    # Calculate vertical positions
    message_y = SCREEN_HEIGHT // 2 - 60
    final_score_y = message_y + message_gap
    save_prompt_y = final_score_y + message_gap
    yes_button_y = save_prompt_y + message_gap
    no_button_y = yes_button_y

    yes_button = pygame.Rect(SCREEN_WIDTH // 2 - 100, yes_button_y, 100, 40)
    no_button = pygame.Rect(SCREEN_WIDTH // 2 + 10, no_button_y, 100, 40)
    yes_text = FONT.render("Yes", True, BLACK)
    no_text = FONT.render("No", True, BLACK)

    screen.blit(game_over_text, (SCREEN_WIDTH // 2 - 100, message_y))
    screen.blit(final_score_text, (SCREEN_WIDTH // 2 - 100, final_score_y))
    screen.blit(save_prompt_text, (SCREEN_WIDTH // 2 - 150, save_prompt_y))
    pygame.draw.rect(screen, BLACK, yes_button, 2)
    pygame.draw.rect(screen, BLACK, no_button, 2)
    screen.blit(yes_text, (yes_button.x + 35, yes_button.y + 10))
    screen.blit(no_text, (no_button.x + 40, no_button.y + 10))
    pygame.display.flip()

    while True:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            elif event.type == pygame.MOUSEBUTTONDOWN:
                mouse_pos = pygame.mouse.get_pos()
                if yes_button.collidepoint(mouse_pos):
                    self.save_to_excel(player_name, self.difficulty, score)
                    return
                elif no_button.collidepoint(mouse_pos):
                    self.display_welcome_screen()
                    return

    def save_to_excel(self, player_name, game_level, score):
        excel_file = "scores.xlsx"

        wb = None
        try:
            wb = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active
        row_num = ws.max_row + 1

        if not ws["A1"].value:
            ws["A1"] = "Player Name"
            ws["B1"] = "Game Level"
            ws["C1"] = "Score"

        ws.cell(row=row_num, column=1, value=player_name)
        ws.cell(row=row_num, column=2, value=game_level)
        ws.cell(row=row_num, column=3, value=score)

        wb.save(excel_file)

    def get_high_score(self, player_name, game_level, current_score):
        excel_file = "scores.xlsx"

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            high_score = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == player_name and row[1] == game_level:
                    high_score = max(high_score, row[2])

            if current_score > high_score:
                return "better"
            elif current_score == high_score:
                return "equal"
            else:
                return "worse"
        except FileNotFoundError:
            return "new"


def setup():
    pygame.init()
    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
    pygame.display.set_caption("Snake Game")
    return screen
