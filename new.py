import pygame
import sys
import os
import random
import openpyxl
from openpyxl import Workbook

# Initialize pygame
pygame.init()

# Constants
SCREEN_WIDTH = 640
SCREEN_HEIGHT = 480
CELL_SIZE = 20
BORDER_SIZE = 20
WHITE = (255, 255, 255)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
BLACK = (0, 0, 0)
FONT = pygame.font.Font(None, 36)

# Set up the screen
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Snake Game")


class Game:
    def __init__(self):
        self.player_name = ""
        self.difficulty = None
        self.running = True
        self.clock = pygame.time.Clock()
        self.screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
        pygame.display.set_caption("Snake Game")

    def display_welcome_screen(self):
        self.screen.fill(WHITE)
        welcome_text = FONT.render("Slithering Challenge", True, BLACK)
        input_rect = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 - 30, 200, 40
        )
        start_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 30, 200, 40
        )
        instruction_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 80, 200, 40
        )
        start_text = FONT.render("Start", True, BLACK)
        instruction_text = FONT.render("Instructions", True, BLACK)

        pygame.draw.rect(self.screen, BLACK, input_rect, 2)
        input_text = FONT.render(self.player_name, True, BLACK)

        # Display placeholder text if player_name is empty
        if not self.player_name:
            input_placeholder = FONT.render("Player Name", True, BLACK)
            self.screen.blit(input_placeholder, (input_rect.x + 5, input_rect.y + 5))
        else:
            self.screen.blit(input_text, (input_rect.x + 5, input_rect.y + 5))

        self.screen.blit(welcome_text, (SCREEN_WIDTH // 2 - 180, SCREEN_HEIGHT // 3))
        pygame.draw.rect(self.screen, BLACK, start_button, 2)
        self.screen.blit(start_text, (start_button.x + 65, start_button.y + 10))
        pygame.draw.rect(self.screen, BLACK, instruction_button, 2)
        self.screen.blit(
            instruction_text, (instruction_button.x + 20, instruction_button.y + 10)
        )

    def display_levels_screen(self):
        self.screen.fill(WHITE)
        levels_text = FONT.render("Levels", True, BLACK)
        easy_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 - 30, 200, 40
        )
        medium_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 30, 200, 40
        )
        hard_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 90, 200, 40
        )
        easy_text = FONT.render("Easy", True, BLACK)
        medium_text = FONT.render("Medium", True, BLACK)
        hard_text = FONT.render("Hard", True, BLACK)

        self.screen.blit(levels_text, (SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 3))
        pygame.draw.rect(self.screen, BLACK, easy_button, 2)
        self.screen.blit(easy_text, (easy_button.x + 70, easy_button.y + 10))
        pygame.draw.rect(self.screen, BLACK, medium_button, 2)
        self.screen.blit(medium_text, (medium_button.x + 45, medium_button.y + 10))
        pygame.draw.rect(self.screen, BLACK, hard_button, 2)
        self.screen.blit(hard_text, (hard_button.x + 70, hard_button.y + 10))

    def game_over(self, screen, player_name, game_level, score):
        self.screen.fill(WHITE)
        game_over_text = FONT.render("Game Over", True, RED)
        final_score_text = FONT.render(f"Final Score: {score}", True, RED)
        player_score_text = FONT.render(f"Player: {self.player_name}", True, RED)
        save_prompt_text = FONT.render("Do you want to save your result?", True, BLACK)
        yes_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 30, 100, 40
        )
        no_button = pygame.Rect(
            SCREEN_WIDTH // 2 + 10, SCREEN_HEIGHT // 2 + 30, 100, 40
        )
        back_button = pygame.Rect(
            SCREEN_WIDTH // 2 - 50, SCREEN_HEIGHT // 2 + 80, 100, 40
        )
        yes_text = FONT.render("Yes", True, BLACK)
        no_text = FONT.render("No", True, BLACK)
        back_text = FONT.render("Back", True, BLACK)

        self.screen.blit(
            game_over_text, (SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 - 30)
        )
        self.screen.blit(
            final_score_text, (SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 10)
        )
        self.screen.blit(
            player_score_text, (SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 50)
        )
        self.screen.blit(
            save_prompt_text, (SCREEN_WIDTH // 2 - 150, SCREEN_HEIGHT // 2 + 100)
        )
        pygame.draw.rect(self.screen, BLACK, yes_button, 2)
        pygame.draw.rect(self.screen, BLACK, no_button, 2)
        pygame.draw.rect(self.screen, BLACK, back_button, 2)
        self.screen.blit(yes_text, (yes_button.x + 35, yes_button.y + 10))
        self.screen.blit(no_text, (no_button.x + 40, no_button.y + 10))
        self.screen.blit(back_text, (back_button.x + 20, back_button.y + 10))
        pygame.display.flip()

        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()
                elif event.type == pygame.MOUSEBUTTONDOWN:
                    mouse_pos = pygame.mouse.get_pos()
                    if yes_button.collidepoint(mouse_pos):
                        # Implement saving score logic here
                        pass
                    elif no_button.collidepoint(mouse_pos):
                        self.display_welcome_screen()
                        return
                    elif back_button.collidepoint(mouse_pos):
                        return

    def save_to_excel(self, player_name, game_level, score):
        excel_file = "scores.xlsx"

        wb = None
        try:
            wb = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active
        row_num = ws.max_row + 1

        if not ws["A1"].value:
            ws["A1"] = "Player Name"
            ws["B1"] = "Game Level"
            ws["C1"] = "Score"

        ws.cell(row=row_num, column=1, value=player_name)
        ws.cell(row=row_num, column=2, value=game_level)
        ws.cell(row=row_num, column=3, value=score)

        wb.save(excel_file)

    def get_high_score(self, player_name, game_level, current_score):
        excel_file = "scores.xlsx"

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            high_score = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == player_name and row[1] == game_level:
                    high_score = max(high_score, row[2])

            if current_score > high_score:
                return "better"
            elif current_score == high_score:
                return "equal"
            else:
                return "worse"
        except FileNotFoundError:
            return "new"

    def run_game(self, game_file):
        pygame.quit()  # Quit the current pygame instance
        os.system(f"python {game_file}")  # Run the selected game file using Python

    def handle_events(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.running = False
            elif event.type == pygame.MOUSEBUTTONDOWN:
                mouse_pos = pygame.mouse.get_pos()
                if self.running and self.difficulty is None:
                    start_button = pygame.Rect(
                        SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 30, 200, 40
                    )
                    instruction_button = pygame.Rect(
                        SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 80, 200, 40
                    )
                    if start_button.collidepoint(mouse_pos):
                        self.difficulty = "easy"  # Default difficulty for demonstration
                elif self.running and self.difficulty is not None:
                    easy_button = pygame.Rect(
                        SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 - 30, 200, 40
                    )
                    medium_button = pygame.Rect(
                        SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 30, 200, 40
                    )
                    hard_button = pygame.Rect(
                        SCREEN_WIDTH // 2 - 100, SCREEN_HEIGHT // 2 + 90, 200, 40
                    )
                    if easy_button.collidepoint(mouse_pos):
                        self.run_game("snake_easy.py")
                    elif medium_button.collidepoint(mouse_pos):
                        self.run_game("snake_medium.py")
                    elif hard_button.collidepoint(mouse_pos):
                        self.run_game("snake_hard.py")

            elif event.type == pygame.KEYDOWN:
                if self.difficulty is None:
                    if event.key == pygame.K_BACKSPACE:
                        self.player_name = self.player_name[:-1]
                    elif event.unicode.isalnum() or event.unicode.isspace():
                        self.player_name += event.unicode

    def run(self):
        while self.running:
            self.handle_events()

            if self.running and self.difficulty is None:
                self.display_welcome_screen()
            elif self.running and self.difficulty is not None:
                self.display_levels_screen()

            pygame.display.flip()
            self.clock.tick(60)

        pygame.quit()
        sys.exit()


class Snake:
    def __init__(self):
        self.positions = [(100, 100), (80, 100), (60, 100)]
        self.direction = (CELL_SIZE, 0)

    def update(self):
        self.positions.insert(
            0,
            (
                self.positions[0][0] + self.direction[0],
                self.positions[0][1] + self.direction[1],
            ),
        )
        self.positions.pop()

    def handle_keys(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_UP and self.direction != (0, CELL_SIZE):
                    self.direction = (0, -CELL_SIZE)
                elif event.key == pygame.K_DOWN and self.direction != (0, -CELL_SIZE):
                    self.direction = (0, CELL_SIZE)
                elif event.key == pygame.K_LEFT and self.direction != (CELL_SIZE, 0):
                    self.direction = (-CELL_SIZE, 0)
                elif event.key == pygame.K_RIGHT and self.direction != (-CELL_SIZE, 0):
                    self.direction = (CELL_SIZE, 0)


class Food:
    def __init__(self, obstacles):
        self.color = RED
        self.obstacles = obstacles
        self.snake_positions = []  # Store snake positions for collision check
        self.update()

    def update(self):
        valid_positions = [
            (x, y)
            for x in range(BORDER_SIZE, SCREEN_WIDTH - BORDER_SIZE, CELL_SIZE)
            for y in range(BORDER_SIZE, SCREEN_HEIGHT - BORDER_SIZE, CELL_SIZE)
            if (x, y) not in [obstacle.position for obstacle in self.obstacles]
            and (x, y) not in self.snake_positions
        ]

        self.position = random.choice(valid_positions)


# Obstacle class
class Obstacle:
    def __init__(self):
        self.color = BLACK
        self.update()

    def update(self):
        valid_positions = [
            (x, y)
            for x in range(BORDER_SIZE, SCREEN_WIDTH - BORDER_SIZE, CELL_SIZE)
            for y in range(BORDER_SIZE, SCREEN_HEIGHT - BORDER_SIZE, CELL_SIZE)
        ]
        self.position = random.choice(valid_positions)


def main():
    game = Game()
    game.run()


if __name__ == "__main__":
    main()
